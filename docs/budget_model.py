"""
Generate EVERSEE Creative Forum budget model (budget_model.xlsx).

Three sheets:
  1. РАСХОДЫ (Expenses) — cost lines with Low/Mid/High and formula-driven totals
  2. ДОХОДЫ (Revenue) — ticket + partner revenue with Early Bird / Regular pricing
  3. АНАЛИЗ (Analysis) — break-even, stress-test matrices, scenario comparison
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
HEADER_FONT = Font(bold=True, size=12, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
SUBHEADER_FONT = Font(bold=True, size=11)
SUBHEADER_FILL = PatternFill("solid", fgColor="D6E4F0")
PARAM_FILL = PatternFill("solid", fgColor="FFFF99")  # yellow — editable
TOTAL_FILL = PatternFill("solid", fgColor="E2EFDA")
TOTAL_FONT = Font(bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
EUR_FMT = '#,##0'
EUR_FMT_NEG = '#,##0;[Red]-#,##0'
PCT_FMT = '0%'


def _style_range(ws, row, col_start, col_end, font=None, fill=None,
                 alignment=None, border=None, number_format=None):
    """Apply styles to a rectangular range in one row."""
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border
        if number_format:
            cell.number_format = number_format


def _set(ws, row, col, value, font=None, fill=None, number_format=None,
         alignment=None, border=None):
    """Write value + optional styles to a single cell."""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if number_format:
        cell.number_format = number_format
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    return cell


# ---------------------------------------------------------------------------
# Sheet 1 — РАСХОДЫ
# ---------------------------------------------------------------------------
def _build_expenses(wb):
    ws = wb.active
    ws.title = "РАСХОДЫ"
    ws.sheet_properties.tabColor = "2F5496"

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 14
    ws.column_dimensions["K"].width = 40

    # -- Parameters block (rows 1-6) ----------------------------------------
    _set(ws, 1, 1, "ПАРАМЕТРЫ", font=HEADER_FONT, fill=HEADER_FILL)
    _style_range(ws, 1, 1, 11, font=HEADER_FONT, fill=HEADER_FILL)

    _set(ws, 2, 1, "Число участников")
    _set(ws, 2, 3, 70, fill=PARAM_FILL, number_format='0')  # C2

    _set(ws, 3, 1, "Доля FULL билетов")
    _set(ws, 3, 3, 0.60, fill=PARAM_FILL, number_format=PCT_FMT)  # C3

    _set(ws, 4, 1, "Плательщик USt (НДС)")
    _set(ws, 4, 3, "Да", fill=PARAM_FILL)  # C4 — "Да" / "Нет"

    # Named references for readability in formulas
    # C2 = participants, C3 = full_share, C4 = USt flag
    P = "'РАСХОДЫ'!$C$2"   # participants
    F = "'РАСХОДЫ'!$C$3"   # FULL share
    FULL = f"ROUND({P}*{F},0)"

    # -- Table header (row 6) -----------------------------------------------
    headers = ["#", "Категория", "Статья", "Ед.", "Low (unit)",
               "Mid (unit)", "High (unit)", "Low (total)", "Mid (total)",
               "High (total)", "Привязка"]
    for i, h in enumerate(headers, 1):
        _set(ws, 6, i, h, font=HEADER_FONT, fill=HEADER_FILL,
             alignment=Alignment(horizontal="center", wrap_text=True))

    # -- Expense lines -------------------------------------------------------
    # Each line: (num, category, item, unit, low, mid, high, formula_type, note)
    # formula_type:
    #   "fix"        — fixed cost, total = unit cost
    #   "x_all"      — per participant, × all × days_multiplier
    #   "x_full"     — per participant FULL
    #   "step_venue"  — stepped venue formula
    #   "step_facil"  — stepped facilitators
    #   "step_bus"    — stepped bus
    #   "pct_rev"     — % of revenue (ticketing commission)
    #   "partial_lin" — partially linear (base + per-person)
    #   "contingency" — % of subtotal
    #   "vat"         — VAT line

    lines = [
        (1,  "Venue", "Аренда Дни 1–2 (зал + breakout + Demo Zone)",
         "2 дня", 8000, 13000, 18000, "step_venue",
         "≤60×0.85, >60×1.0, >85×1.15"),
        (2,  "Venue", "Аренда День 3 (утренняя сессия, без breakout)",
         "1 день", 2000, 3500, 5000, "fix", "Фикс."),
        (3,  "Кейтеринг", "Кофе-брейки + обеды, Дни 1–2",
         "чел/день", 35, 55, 75, "x_all_2",
         "× участников × 2 дня"),
        (4,  "Кейтеринг", "Кофе + лёгкий обед, День 3",
         "чел", 20, 30, 40, "x_full", "× участников FULL"),
        (5,  "Ужины", "Групповой ужин #1 (чт, рядом с Alizée)",
         "чел", 60, 90, 130, "x_full", "× участников FULL"),
        (6,  "Ужины", "Групповой ужин #2 (пт, рядом с Palast)",
         "чел", 60, 90, 130, "x_full", "× участников FULL"),
        (7,  "Шоу", "Билеты Alizée (групповая скидка)",
         "чел", 53, 75, 94, "x_full", "× участников FULL"),
        (8,  "Шоу", "Билеты Friedrichstadt-Palast",
         "чел", 40, 60, 80, "x_full", "× участников FULL"),
        (9,  "AV", "Оборудование + техник (2 дня)",
         "комплект", 2500, 5000, 8000, "fix", "Фикс."),
        (10, "Фасилитаторы", "Главный ведущий (2 дня + подготовка)",
         "чел", 3000, 5000, 8000, "fix", "Фикс."),
        (11, "Фасилитаторы", "Breakout-фасилитаторы (5–6, День 2)",
         "день", 5000, 8000, 12000, "step_facil",
         "Число групп зависит от участников"),
        (12, "Фото/видео", "Фотограф (2 дня)",
         "комплект", 1800, 2500, 4000, "fix", "Фикс."),
        (13, "Фото/видео", "Видеограф + монтаж",
         "комплект", 2000, 4000, 7000, "fix", "Фикс."),
        (14, "Печать", "Бейджи, карточки, баннеры, материалы",
         "комплект", 1500, 2500, 4000, "partial_lin",
         "Частично лин. (бейджи/карточки)"),
        (15, "Транспорт", "Автобус(ы) трансферы, 2 вечера",
         "комплект", 1500, 3000, 5000, "step_bus",
         "1 бус ≤50, 2 буса >50"),
        (16, "Страховка", "Ответственность + отмена",
         "комплект", 200, 500, 1000, "fix", "Фикс."),
        (17, "Сайт/маркетинг", "Лендинг, рассылки, LinkedIn, дизайн",
         "комплект", 1500, 3000, 5000, "fix", "Фикс."),
        (18, "Команда орг.", "Перелёты + проживание команды (3–4 чел, 4–5 ночей)",
         "комплект", 3000, 5000, 8000, "fix", "Фикс."),
        (19, "Тикетинг", "Комиссия платформы (Eventbrite/Stripe)",
         "% от выручки", 0.03, 0.04, 0.05, "pct_rev",
         "% от общей выручки с билетов"),
        (20, "Юрист/бух", "Контракты, GDPR, налоговый учёт",
         "комплект", 1000, 2000, 3500, "fix", "Фикс."),
    ]

    start_row = 7
    data_rows = {}  # num -> row for reference

    for idx, line in enumerate(lines):
        r = start_row + idx
        num, cat, item, unit, low, mid, high, ftype, note = line
        data_rows[num] = r

        _set(ws, r, 1, num, border=THIN_BORDER,
             alignment=Alignment(horizontal="center"))
        _set(ws, r, 2, cat, border=THIN_BORDER)
        _set(ws, r, 3, item, border=THIN_BORDER)
        _set(ws, r, 4, unit, border=THIN_BORDER,
             alignment=Alignment(horizontal="center"))

        # Unit costs (E, F, G)
        fmt = PCT_FMT if ftype == "pct_rev" else EUR_FMT
        _set(ws, r, 5, low, border=THIN_BORDER, number_format=fmt)
        _set(ws, r, 6, mid, border=THIN_BORDER, number_format=fmt)
        _set(ws, r, 7, high, border=THIN_BORDER, number_format=fmt)

        # Total formulas (H, I, J) — one for each scenario
        for col_offset, unit_col in enumerate(["E", "F", "G"]):
            tcol = 8 + col_offset  # H=8, I=9, J=10
            uc = f"{unit_col}{r}"  # unit-cost cell ref

            if ftype == "fix":
                formula = f"={uc}"
            elif ftype == "x_all_2":
                formula = f"={uc}*{P}*2"
            elif ftype == "x_full":
                formula = f"={uc}*{FULL}"
            elif ftype == "step_venue":
                # ≤60 → ×0.85, 61-85 → ×1.0, >85 → ×1.15
                formula = (
                    f"=IF({P}<=60,{uc}*0.85,"
                    f"IF({P}<=85,{uc},"
                    f"{uc}*1.15))"
                )
            elif ftype == "step_facil":
                # ≤50 → 5 groups (×1.0), 51-75 → 6 groups (×1.2), >75 → 7 (×1.4)
                formula = (
                    f"=IF({P}<=50,{uc},"
                    f"IF({P}<=75,{uc}*1.2,"
                    f"{uc}*1.4))"
                )
            elif ftype == "step_bus":
                # ≤50 → ×1, >50 → ×2
                formula = f"=IF({P}<=50,{uc},{uc}*2)"
            elif ftype == "partial_lin":
                # 70% fixed base + 30% scales with participants (base=70 ppl)
                formula = f"={uc}*0.7+{uc}*0.3*{P}/70"
            elif ftype == "pct_rev":
                # Will reference total ticket revenue from sheet 2
                # Revenue total is on ДОХОДЫ sheet — we'll place it in a known cell
                rev_ref = "'ДОХОДЫ'!$E$30"
                formula = f"={uc}*{rev_ref}"
            else:
                formula = f"={uc}"

            _set(ws, r, tcol, formula, border=THIN_BORDER,
                 number_format=EUR_FMT)

        _set(ws, r, 11, note, border=THIN_BORDER,
             font=Font(italic=True, color="666666"))

    # -- Subtotal rows 1-20 --------------------------------------------------
    sub_row = start_row + len(lines)  # row 27
    _set(ws, sub_row, 1, "", fill=SUBHEADER_FILL)
    _set(ws, sub_row, 2, "", fill=SUBHEADER_FILL)
    _set(ws, sub_row, 3, "Промежуточный итог (стр. 1–20)",
         font=TOTAL_FONT, fill=SUBHEADER_FILL)
    for c in range(4, 12):
        ws.cell(row=sub_row, column=c).fill = SUBHEADER_FILL
    for tcol_letter, tcol in [("H", 8), ("I", 9), ("J", 10)]:
        first = start_row
        last = sub_row - 1
        formula = f"=SUM({tcol_letter}{first}:{tcol_letter}{last})"
        _set(ws, sub_row, tcol, formula, font=TOTAL_FONT,
             fill=SUBHEADER_FILL, border=THIN_BORDER, number_format=EUR_FMT)

    # -- Row 21: НДС --------------------------------------------------------
    vat_row = sub_row + 1
    _set(ws, vat_row, 1, 21, border=THIN_BORDER,
         alignment=Alignment(horizontal="center"))
    _set(ws, vat_row, 2, "НДС (USt)", border=THIN_BORDER)
    _set(ws, vat_row, 3, "19% на расходы (если не плательщик USt)",
         border=THIN_BORDER)
    _set(ws, vat_row, 4, "%", border=THIN_BORDER,
         alignment=Alignment(horizontal="center"))
    for unit_col in [5, 6, 7]:
        _set(ws, vat_row, unit_col, 0.19, border=THIN_BORDER,
             number_format=PCT_FMT)
    # If USt payer (C4="Да") → 0, else ~15% of subtotal (not everything is taxable)
    for tcol_letter, tcol in [("H", 8), ("I", 9), ("J", 10)]:
        formula = (
            f'=IF($C$4="Да",0,'
            f'{tcol_letter}{sub_row}*0.15)'
        )
        _set(ws, vat_row, tcol, formula, border=THIN_BORDER,
             number_format=EUR_FMT)
    _set(ws, vat_row, 11,
         "Если плательщик USt — Vorsteuer вычитается, строка = 0",
         border=THIN_BORDER, font=Font(italic=True, color="666666"))
    data_rows[21] = vat_row

    # -- Row 22: Contingency -------------------------------------------------
    cont_row = vat_row + 1
    _set(ws, cont_row, 1, 22, border=THIN_BORDER,
         alignment=Alignment(horizontal="center"))
    _set(ws, cont_row, 2, "Прочее", border=THIN_BORDER)
    _set(ws, cont_row, 3, "Непредвиденные расходы", border=THIN_BORDER)
    _set(ws, cont_row, 4, "%", border=THIN_BORDER,
         alignment=Alignment(horizontal="center"))
    cont_rates = [0.08, 0.10, 0.12]
    for i, rate in enumerate(cont_rates):
        _set(ws, cont_row, 5 + i, rate, border=THIN_BORDER,
             number_format=PCT_FMT)
    for i, (tcol_letter, ucol_letter) in enumerate(
            [("H", "E"), ("I", "F"), ("J", "G")]):
        formula = f"={ucol_letter}{cont_row}*{tcol_letter}{sub_row}"
        _set(ws, cont_row, 8 + i, formula, border=THIN_BORDER,
             number_format=EUR_FMT)
    _set(ws, cont_row, 11, "% от суммы строк 1–20",
         border=THIN_BORDER, font=Font(italic=True, color="666666"))
    data_rows[22] = cont_row

    # -- TOTAL row -----------------------------------------------------------
    total_row = cont_row + 2
    _set(ws, total_row, 3, "ИТОГО РАСХОДЫ", font=Font(bold=True, size=13),
         fill=TOTAL_FILL)
    _style_range(ws, total_row, 1, 11, fill=TOTAL_FILL)
    for tcol_letter, tcol in [("H", 8), ("I", 9), ("J", 10)]:
        formula = (
            f"={tcol_letter}{sub_row}"
            f"+{tcol_letter}{vat_row}"
            f"+{tcol_letter}{cont_row}"
        )
        _set(ws, total_row, tcol, formula, font=Font(bold=True, size=13),
             fill=TOTAL_FILL, border=THIN_BORDER, number_format=EUR_FMT)

    # -- Note about НДС ------------------------------------------------------
    note_row = total_row + 2
    _set(ws, note_row, 1,
         "Примечание по НДС: Если EVERSEE — плательщик USt, расходы netto, "
         "доходы netto, строка НДС = 0. Если нет — НДС ~15% от суммы расходов.",
         font=Font(italic=True, color="666666"))
    ws.merge_cells(start_row=note_row, start_column=1,
                   end_row=note_row, end_column=11)

    return {
        "participants": "$C$2",
        "full_share": "$C$3",
        "ust_flag": "$C$4",
        "sub_row": sub_row,
        "vat_row": vat_row,
        "cont_row": cont_row,
        "total_row": total_row,
        "start_row": start_row,
        "data_rows": data_rows,
    }


# ---------------------------------------------------------------------------
# Sheet 2 — ДОХОДЫ
# ---------------------------------------------------------------------------
def _build_revenue(wb, exp_info):
    ws = wb.create_sheet("ДОХОДЫ")
    ws.sheet_properties.tabColor = "548235"

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 18

    P = "'РАСХОДЫ'!$C$2"   # participants
    FS = "'РАСХОДЫ'!$C$3"  # FULL share

    # -- Parameters (rows 1-11) ---------------------------------------------
    _set(ws, 1, 1, "ПАРАМЕТРЫ ДОХОДОВ", font=HEADER_FONT, fill=HEADER_FILL)
    _style_range(ws, 1, 1, 5, font=HEADER_FONT, fill=HEADER_FILL)

    params = [
        (2, "Доля Creative / Tech", "creative_share", 0.60, PCT_FMT,
         "60% Creative, 40% Tech"),
        (3, "Доля Early Bird", "eb_share", 0.50, PCT_FMT, ""),
        (4, "Доля FULL vs FORUM", "full_share_rev", None, PCT_FMT,
         "Связана с РАСХОДЫ!C3"),
        (5, "Слоты Lightning Talk (продано)", "lt_sold", 8, '0', "из 12"),
        (6, "Слоты Keynote (продано)", "kn_sold", 1, '0', "из 2"),
        (7, "Partner пакетов", "partner_n", 1, '0', "вкл. 2 билета TECH FULL"),
        (8, "Strategic Partner пакетов", "strategic_n", 1, '0',
         "вкл. 4 билета TECH FULL"),
        (9, "Title Partner пакетов", "title_n", 0, '0',
         "вкл. 6 билетов TECH FULL"),
    ]
    for r, label, _name, val, fmt, note in params:
        _set(ws, r, 1, label)
        if val is not None:
            _set(ws, r, 3, val, fill=PARAM_FILL, number_format=fmt)
        else:
            # Link to expenses sheet FULL share
            _set(ws, r, 3, f"={FS}", fill=PatternFill("solid", fgColor="E2EFDA"),
                 number_format=fmt)
        if note:
            _set(ws, r, 5, note, font=Font(italic=True, color="666666"))

    # Derived: partner tickets to subtract from paid TECH FULL
    # Row 10: total partner tickets
    _set(ws, 10, 1, "Билеты из пакетов (TECH FULL)")
    _set(ws, 10, 3, "=$C$7*2+$C$8*4+$C$9*6",
         number_format='0')
    _set(ws, 10, 5, "Вычитаются из платных TECH билетов",
         font=Font(italic=True, color="666666"))

    # -- Revenue table header (row 12) --------------------------------------
    _set(ws, 12, 1, "ДОХОДЫ ОТ БИЛЕТОВ И ПАРТНЁРОВ",
         font=HEADER_FONT, fill=HEADER_FILL)
    _style_range(ws, 12, 1, 5, font=HEADER_FONT, fill=HEADER_FILL)

    cols = ["Источник", "Кол-во (расчёт)", "Цена Early Bird",
            "Цена Regular (+30%)", "Итого"]
    for i, c in enumerate(cols, 1):
        _set(ws, 13, i, c, font=Font(bold=True, color="FFFFFF"),
             fill=PatternFill("solid", fgColor="548235"),
             alignment=Alignment(horizontal="center", wrap_text=True),
             border=THIN_BORDER)

    # Ticket types:
    # Creative share = C2, Tech = 1-C2
    # FORUM share = 1 - full_share (C4), FULL share = C4
    # EB share = C3, Regular = 1 - C3
    # Quantity = participants × track_share × type_share
    #   EB qty = total × eb_share, Regular qty = total × (1-eb_share)
    # Revenue = EB_qty × EB_price + Reg_qty × Reg_price

    # Prices (EB → Regular = EB × 1.3)
    ticket_lines = [
        # (label, qty_formula, eb_price)
        ("CREATIVE FORUM",
         f"ROUND({P}*$C$2*(1-$C$4),0)",
         290),
        ("CREATIVE FULL",
         f"ROUND({P}*$C$2*$C$4,0)",
         590),
        ("TECH FORUM",
         f"MAX(ROUND({P}*(1-$C$2)*(1-$C$4),0)-0,0)",
         490),
        ("TECH FULL",
         f"MAX(ROUND({P}*(1-$C$2)*$C$4,0)-$C$10,0)",
         890),
    ]

    row = 14
    ticket_total_cells = []

    for label, qty_formula, eb_price in ticket_lines:
        _set(ws, row, 1, label, border=THIN_BORDER, font=Font(bold=True))
        _set(ws, row, 2, f"={qty_formula}", border=THIN_BORDER,
             number_format='0')
        _set(ws, row, 3, eb_price, border=THIN_BORDER, number_format=EUR_FMT)
        reg_price = round(eb_price * 1.3)
        _set(ws, row, 4, reg_price, border=THIN_BORDER, number_format=EUR_FMT)
        # Revenue = qty × (eb_share × eb_price + (1-eb_share) × reg_price)
        revenue_formula = (
            f"=B{row}*($C$3*C{row}+(1-$C$3)*D{row})"
        )
        _set(ws, row, 5, revenue_formula, border=THIN_BORDER,
             number_format=EUR_FMT)
        ticket_total_cells.append(f"E{row}")
        row += 1

    # Spacer
    row += 1

    # Speaking slots
    _set(ws, row, 1, "СЛОТЫ ВЫСТУПЛЕНИЙ", font=SUBHEADER_FONT,
         fill=SUBHEADER_FILL)
    _style_range(ws, row, 1, 5, fill=SUBHEADER_FILL)
    row += 1

    lt_row = row
    _set(ws, row, 1, "Lightning Talk", border=THIN_BORDER, font=Font(bold=True))
    _set(ws, row, 2, "=$C$5", border=THIN_BORDER, number_format='0')
    _set(ws, row, 3, 500, border=THIN_BORDER, number_format=EUR_FMT)
    _set(ws, row, 4, "", border=THIN_BORDER)
    _set(ws, row, 5, f"=B{row}*C{row}", border=THIN_BORDER,
         number_format=EUR_FMT)
    ticket_total_cells.append(f"E{row}")
    row += 1

    kn_row = row
    _set(ws, row, 1, "Keynote", border=THIN_BORDER, font=Font(bold=True))
    _set(ws, row, 2, "=$C$6", border=THIN_BORDER, number_format='0')
    _set(ws, row, 3, 1500, border=THIN_BORDER, number_format=EUR_FMT)
    _set(ws, row, 4, "", border=THIN_BORDER)
    _set(ws, row, 5, f"=B{row}*C{row}", border=THIN_BORDER,
         number_format=EUR_FMT)
    ticket_total_cells.append(f"E{row}")
    row += 1

    # Spacer
    row += 1

    # Partner packages
    _set(ws, row, 1, "ПАРТНЁРСКИЕ ПАКЕТЫ", font=SUBHEADER_FONT,
         fill=SUBHEADER_FILL)
    _style_range(ws, row, 1, 5, fill=SUBHEADER_FILL)
    row += 1

    partner_lines = [
        ("Partner", "=$C$7", 3000),
        ("Strategic Partner", "=$C$8", 8000),
        ("Title Partner", "=$C$9", 15000),
    ]
    for label, qty_formula, price in partner_lines:
        _set(ws, row, 1, label, border=THIN_BORDER, font=Font(bold=True))
        _set(ws, row, 2, qty_formula, border=THIN_BORDER, number_format='0')
        _set(ws, row, 3, price, border=THIN_BORDER, number_format=EUR_FMT)
        _set(ws, row, 4, "", border=THIN_BORDER)
        _set(ws, row, 5, f"=B{row}*C{row}", border=THIN_BORDER,
             number_format=EUR_FMT)
        ticket_total_cells.append(f"E{row}")
        row += 1

    # -- TOTAL revenue -------------------------------------------------------
    row += 1
    total_rev_row = row
    _set(ws, row, 1, "ИТОГО ДОХОДЫ", font=Font(bold=True, size=13),
         fill=TOTAL_FILL)
    _style_range(ws, row, 1, 5, fill=TOTAL_FILL)
    formula = "=" + "+".join(ticket_total_cells)
    _set(ws, row, 5, formula, font=Font(bold=True, size=13),
         fill=TOTAL_FILL, border=THIN_BORDER, number_format=EUR_FMT)

    # Ensure the total is in a predictable cell for cross-sheet references.
    # The plan says E30 is used by ticketing commission — let's place total there.
    # If total_rev_row != 30, we add a helper cell at row 30
    if total_rev_row != 30:
        _set(ws, 30, 1, "Выручка (ссылка для расчётов)")
        _set(ws, 30, 5, f"=E{total_rev_row}", number_format=EUR_FMT)

    return {
        "total_rev_row": total_rev_row,
        "ticket_total_cells": ticket_total_cells,
    }


# ---------------------------------------------------------------------------
# Sheet 3 — АНАЛИЗ
# ---------------------------------------------------------------------------
def _build_analysis(wb, exp_info, rev_info):
    ws = wb.create_sheet("АНАЛИЗ")
    ws.sheet_properties.tabColor = "BF8F00"

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 16

    P = "'РАСХОДЫ'!$C$2"
    tr = exp_info["total_row"]
    sub = exp_info["sub_row"]

    # ========== Block 1: Base scenario summary (rows 1-8) ==================
    _set(ws, 1, 1, "СВОДКА БАЗОВОГО СЦЕНАРИЯ (Mid)",
         font=HEADER_FONT, fill=HEADER_FILL)
    _style_range(ws, 1, 1, 4, font=HEADER_FONT, fill=HEADER_FILL)

    _set(ws, 3, 1, "Итого расходы (Mid)", font=Font(bold=True))
    _set(ws, 3, 2, f"='РАСХОДЫ'!I{tr}", number_format=EUR_FMT)

    _set(ws, 4, 1, "Итого доходы", font=Font(bold=True))
    _set(ws, 4, 2, "='ДОХОДЫ'!E30", number_format=EUR_FMT)

    _set(ws, 5, 1, "Комиссия тикетинга (Mid)", font=Font(bold=True))
    comm_row = exp_info["data_rows"][19]
    _set(ws, 5, 2, f"='РАСХОДЫ'!I{comm_row}", number_format=EUR_FMT)

    _set(ws, 6, 1, "Чистый результат", font=Font(bold=True, size=12))
    _set(ws, 6, 2, "=B4-B3", number_format=EUR_FMT_NEG,
         font=Font(bold=True, size=12))

    _set(ws, 7, 1, "Маржа", font=Font(bold=True))
    _set(ws, 7, 2, "=IF(B4=0,0,B6/B4)", number_format='0.0%')

    # ========== Block 2: Break-even table (rows 10-28) =====================
    _set(ws, 10, 1, "BREAK-EVEN АНАЛИЗ",
         font=HEADER_FONT, fill=HEADER_FILL)
    _style_range(ws, 10, 1, 5, font=HEADER_FONT, fill=HEADER_FILL)

    be_headers = ["Участников", "Расходы (Mid)", "Доходы", "Результат", "Маржа"]
    for i, h in enumerate(be_headers, 1):
        _set(ws, 11, i, h, font=Font(bold=True, color="FFFFFF"),
             fill=PatternFill("solid", fgColor="BF8F00"),
             border=THIN_BORDER,
             alignment=Alignment(horizontal="center"))

    # We can't dynamically recalculate with different participant counts
    # in pure Excel without VBA or data tables. Instead, we build static
    # formulas that replicate the model logic for each N.

    def _expense_formula_for_n(n, col="I"):
        """Build a simplified Mid expense total formula for a given N.

        Replicates the key variable-cost logic without circular refs.
        """
        full = f"ROUND({n}*'РАСХОДЫ'!$C$3,0)"

        # Venue days 1-2 (step)
        venue12 = f"IF({n}<=60,13000*0.85,IF({n}<=85,13000,13000*1.15))"
        venue3 = "3500"
        # Catering days 1-2
        cater12 = f"55*{n}*2"
        # Catering day 3 (FULL)
        cater3 = f"30*{full}"
        # Dinners
        dinner1 = f"90*{full}"
        dinner2 = f"90*{full}"
        # Shows
        show1 = f"75*{full}"
        show2 = f"60*{full}"
        # Fixed costs (AV + facilitators + photo + video + print + bus +
        #   insurance + marketing + team + legal)
        av = "5000"
        main_host = "5000"
        facil = f"IF({n}<=50,8000,IF({n}<=75,8000*1.2,8000*1.4))"
        photo = "2500"
        video = "4000"
        printing = f"2500*0.7+2500*0.3*{n}/70"
        bus = f"IF({n}<=50,3000,3000*2)"
        insurance = "500"
        marketing = "3000"
        team = "5000"
        legal = "2000"

        subtotal = (
            f"({venue12})+{venue3}+{cater12}+{cater3}"
            f"+{dinner1}+{dinner2}+{show1}+{show2}"
            f"+{av}+{main_host}+({facil})+{photo}+{video}"
            f"+({printing})+({bus})+{insurance}+{marketing}"
            f"+{team}+{legal}"
        )
        # Ticketing commission (4% of revenue) — add later
        # VAT
        vat = f'IF(\'РАСХОДЫ\'!$C$4="Да",0,({subtotal})*0.15)'
        # Contingency 10%
        contingency = f"({subtotal})*0.1"

        return f"({subtotal})+({vat})+({contingency})"

    def _revenue_formula_for_n(n):
        """Build a simplified revenue formula for a given N.

        Uses current parameter values from ДОХОДЫ sheet.
        """
        cs = "'ДОХОДЫ'!$C$2"   # creative share
        eb = "'ДОХОДЫ'!$C$3"   # EB share
        fs = "'РАСХОДЫ'!$C$3"  # FULL share
        partner_tix = "'ДОХОДЫ'!$C$10"  # partner tickets

        # Quantities
        cr_forum = f"ROUND({n}*{cs}*(1-{fs}),0)"
        cr_full = f"ROUND({n}*{cs}*{fs},0)"
        tech_forum = f"MAX(ROUND({n}*(1-{cs})*(1-{fs}),0),0)"
        tech_full = f"MAX(ROUND({n}*(1-{cs})*{fs},0)-{partner_tix},0)"

        # Revenue per type = qty × (eb_share × eb_price + (1-eb) × reg_price)
        cr_forum_rev = f"({cr_forum})*({eb}*290+(1-{eb})*377)"
        cr_full_rev = f"({cr_full})*({eb}*590+(1-{eb})*767)"
        tech_forum_rev = f"({tech_forum})*({eb}*490+(1-{eb})*637)"
        tech_full_rev = f"({tech_full})*({eb}*890+(1-{eb})*1157)"

        # Speaking slots
        lt = "'ДОХОДЫ'!$C$5*500"
        kn = "'ДОХОДЫ'!$C$6*1500"

        # Partners
        partners = ("'ДОХОДЫ'!$C$7*3000+'ДОХОДЫ'!$C$8*8000"
                     "+'ДОХОДЫ'!$C$9*15000")

        return (f"{cr_forum_rev}+{cr_full_rev}"
                f"+{tech_forum_rev}+{tech_full_rev}"
                f"+{lt}+{kn}+{partners}")

    be_row = 12
    participants_range = list(range(30, 105, 5))
    for n in participants_range:
        _set(ws, be_row, 1, n, border=THIN_BORDER,
             alignment=Alignment(horizontal="center"))

        exp_f = _expense_formula_for_n(n)
        rev_f = _revenue_formula_for_n(n)

        # Also add ticketing commission to expenses
        full_exp = f"={exp_f}+0.04*({rev_f})"
        _set(ws, be_row, 2, full_exp, border=THIN_BORDER,
             number_format=EUR_FMT)

        _set(ws, be_row, 3, f"={rev_f}", border=THIN_BORDER,
             number_format=EUR_FMT)

        _set(ws, be_row, 4, f"=C{be_row}-B{be_row}", border=THIN_BORDER,
             number_format=EUR_FMT_NEG)

        _set(ws, be_row, 5, f"=IF(C{be_row}=0,0,D{be_row}/C{be_row})",
             border=THIN_BORDER, number_format='0.0%')
        be_row += 1

    # Conditional formatting for result column D
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    ws.conditional_formatting.add(
        f"D12:D{be_row - 1}",
        CellIsRule(operator="greaterThanOrEqual", formula=["0"],
                   fill=green_fill))
    ws.conditional_formatting.add(
        f"D12:D{be_row - 1}",
        CellIsRule(operator="lessThan", formula=["0"], fill=red_fill))

    # ========== Block 3: Matrix Participants × Partners (rows be_row+2..) ===
    mat1_start = be_row + 2
    _set(ws, mat1_start, 1, "МАТРИЦА: Участники × Партнёры",
         font=HEADER_FONT, fill=HEADER_FILL)
    _style_range(ws, mat1_start, 1, 6, font=HEADER_FONT, fill=HEADER_FILL)

    partner_scenarios = [
        ("0 партн.", 0, 0, 0),
        ("1 Partner", 1, 0, 0),
        ("1P + 1S", 1, 1, 0),
        ("1S + 1T", 0, 1, 1),
        ("1P+1S+1T", 1, 1, 1),
    ]
    # Header row
    hr = mat1_start + 1
    _set(ws, hr, 1, "Участников", font=Font(bold=True),
         fill=PatternFill("solid", fgColor="BF8F00"), border=THIN_BORDER)
    for i, (label, _, _, _) in enumerate(partner_scenarios):
        _set(ws, hr, 2 + i, label, font=Font(bold=True),
             fill=PatternFill("solid", fgColor="BF8F00"), border=THIN_BORDER,
             alignment=Alignment(horizontal="center"))

    def _revenue_formula_for_n_partners(n, p_count, s_count, t_count):
        """Revenue formula with specific partner configuration."""
        cs = "'ДОХОДЫ'!$C$2"
        eb = "'ДОХОДЫ'!$C$3"
        fs = "'РАСХОДЫ'!$C$3"
        partner_tix = f"({p_count}*2+{s_count}*4+{t_count}*6)"

        cr_forum = f"ROUND({n}*{cs}*(1-{fs}),0)"
        cr_full = f"ROUND({n}*{cs}*{fs},0)"
        tech_forum = f"MAX(ROUND({n}*(1-{cs})*(1-{fs}),0),0)"
        tech_full = f"MAX(ROUND({n}*(1-{cs})*{fs},0)-{partner_tix},0)"

        cr_forum_rev = f"({cr_forum})*({eb}*290+(1-{eb})*377)"
        cr_full_rev = f"({cr_full})*({eb}*590+(1-{eb})*767)"
        tech_forum_rev = f"({tech_forum})*({eb}*490+(1-{eb})*637)"
        tech_full_rev = f"({tech_full})*({eb}*890+(1-{eb})*1157)"

        lt = "'ДОХОДЫ'!$C$5*500"
        kn = "'ДОХОДЫ'!$C$6*1500"
        partners = f"{p_count}*3000+{s_count}*8000+{t_count}*15000"

        return (f"{cr_forum_rev}+{cr_full_rev}"
                f"+{tech_forum_rev}+{tech_full_rev}"
                f"+{lt}+{kn}+{partners}")

    participants_list = [30, 40, 50, 60, 70, 80, 90, 100]
    data_start = hr + 1
    for row_idx, n in enumerate(participants_list):
        r = data_start + row_idx
        _set(ws, r, 1, n, border=THIN_BORDER,
             alignment=Alignment(horizontal="center"), font=Font(bold=True))
        for col_idx, (_, p, s, t) in enumerate(partner_scenarios):
            rev = _revenue_formula_for_n_partners(n, p, s, t)
            exp = _expense_formula_for_n(n)
            formula = f"=({rev})-({exp})-0.04*({rev})"
            _set(ws, r, 2 + col_idx, formula, border=THIN_BORDER,
                 number_format=EUR_FMT_NEG)
    mat1_end = data_start + len(participants_list) - 1

    # Conditional formatting
    data_range = f"B{data_start}:F{mat1_end}"
    ws.conditional_formatting.add(
        data_range,
        CellIsRule(operator="greaterThanOrEqual", formula=["0"],
                   fill=green_fill))
    ws.conditional_formatting.add(
        data_range,
        CellIsRule(operator="lessThan", formula=["0"], fill=red_fill))

    # ========== Block 4: Matrix Participants × Pricing (rows after mat1) ===
    mat2_start = mat1_end + 3
    _set(ws, mat2_start, 1, "МАТРИЦА: Участники × Ценовой сценарий",
         font=HEADER_FONT, fill=HEADER_FILL)
    _style_range(ws, mat2_start, 1, 5, font=HEADER_FONT, fill=HEADER_FILL)

    price_scenarios = [
        ("EB −20%", 0.8),
        ("Early Bird", 1.0),
        ("Regular", 1.3),
        ("Regular +20%", 1.3 * 1.2),
    ]
    hr2 = mat2_start + 1
    _set(ws, hr2, 1, "Участников", font=Font(bold=True),
         fill=PatternFill("solid", fgColor="BF8F00"), border=THIN_BORDER)
    for i, (label, _) in enumerate(price_scenarios):
        _set(ws, hr2, 2 + i, label, font=Font(bold=True),
             fill=PatternFill("solid", fgColor="BF8F00"), border=THIN_BORDER,
             alignment=Alignment(horizontal="center"))

    def _revenue_formula_for_n_price(n, price_mult):
        """Revenue with price multiplier (base = EB prices). Partners: 1P+1S."""
        cs = "'ДОХОДЫ'!$C$2"
        fs = "'РАСХОДЫ'!$C$3"
        partner_tix = f"(1*2+1*4+0*6)"  # 1P+1S fixed

        cr_forum = f"ROUND({n}*{cs}*(1-{fs}),0)"
        cr_full = f"ROUND({n}*{cs}*{fs},0)"
        tech_forum = f"MAX(ROUND({n}*(1-{cs})*(1-{fs}),0),0)"
        tech_full = f"MAX(ROUND({n}*(1-{cs})*{fs},0)-{partner_tix},0)"

        m = price_mult
        cr_forum_rev = f"({cr_forum})*{290 * m:.0f}"
        cr_full_rev = f"({cr_full})*{590 * m:.0f}"
        tech_forum_rev = f"({tech_forum})*{490 * m:.0f}"
        tech_full_rev = f"({tech_full})*{890 * m:.0f}"

        lt = f"'ДОХОДЫ'!$C$5*{500 * m:.0f}"
        kn = f"'ДОХОДЫ'!$C$6*{1500 * m:.0f}"
        partners = "1*3000+1*8000+0*15000"  # partners fixed

        return (f"{cr_forum_rev}+{cr_full_rev}"
                f"+{tech_forum_rev}+{tech_full_rev}"
                f"+{lt}+{kn}+{partners}")

    data2_start = hr2 + 1
    for row_idx, n in enumerate(participants_list):
        r = data2_start + row_idx
        _set(ws, r, 1, n, border=THIN_BORDER,
             alignment=Alignment(horizontal="center"), font=Font(bold=True))
        for col_idx, (_, mult) in enumerate(price_scenarios):
            rev = _revenue_formula_for_n_price(n, mult)
            exp = _expense_formula_for_n(n)
            formula = f"=({rev})-({exp})-0.04*({rev})"
            _set(ws, r, 2 + col_idx, formula, border=THIN_BORDER,
                 number_format=EUR_FMT_NEG)
    mat2_end = data2_start + len(participants_list) - 1

    data_range2 = f"B{data2_start}:E{mat2_end}"
    ws.conditional_formatting.add(
        data_range2,
        CellIsRule(operator="greaterThanOrEqual", formula=["0"],
                   fill=green_fill))
    ws.conditional_formatting.add(
        data_range2,
        CellIsRule(operator="lessThan", formula=["0"], fill=red_fill))

    # ========== Block 5: Three scenarios (rows after mat2) =================
    sc_start = mat2_end + 3
    _set(ws, sc_start, 1, "ТРИ СЦЕНАРИЯ",
         font=HEADER_FONT, fill=HEADER_FILL)
    _style_range(ws, sc_start, 1, 4, font=HEADER_FONT, fill=HEADER_FILL)

    sc_headers = ["Параметр", "Пессимист", "Базовый", "Оптимист"]
    hr3 = sc_start + 1
    for i, h in enumerate(sc_headers, 1):
        _set(ws, hr3, i, h, font=Font(bold=True, color="FFFFFF"),
             fill=PatternFill("solid", fgColor="BF8F00"), border=THIN_BORDER,
             alignment=Alignment(horizontal="center"))

    scenario_params = [
        ("Участников", 40, 70, 100),
        ("Расходы (сценарий)", "High", "Mid", "Low"),
        ("Доля Early Bird", "80%", "50%", "30%"),
        ("Партнёры", "0", "1P + 1S", "1S + 1T"),
        ("Доля FULL", "40%", "60%", "70%"),
        ("Слоты Lightning", 4, 8, 12),
        ("Слоты Keynote", 0, 1, 2),
    ]
    for idx, (param, pess, base, opt) in enumerate(scenario_params):
        r = hr3 + 1 + idx
        _set(ws, r, 1, param, border=THIN_BORDER, font=Font(bold=True))
        _set(ws, r, 2, pess, border=THIN_BORDER,
             alignment=Alignment(horizontal="center"))
        _set(ws, r, 3, base, border=THIN_BORDER,
             alignment=Alignment(horizontal="center"))
        _set(ws, r, 4, opt, border=THIN_BORDER,
             alignment=Alignment(horizontal="center"))

    # Computed results for each scenario
    res_start = hr3 + 1 + len(scenario_params) + 1

    def _scenario_expense(n, level, full_share, lt, kn, p, s, t):
        """Build expense formula for a specific scenario."""
        full = f"ROUND({n}*{full_share},0)"

        if level == "Low":
            v12, v3 = 8000, 2000
            cat12, cat3 = 35, 20
            d1, d2 = 60, 60
            s1, s2 = 53, 40
            av, host = 2500, 3000
            facil_base = 5000
            photo, video = 1800, 2000
            print_base = 1500
            bus_base = 1500
            ins, mkt, team, legal = 200, 1500, 3000, 1000
            cont_rate = 0.08
        elif level == "High":
            v12, v3 = 18000, 5000
            cat12, cat3 = 75, 40
            d1, d2 = 130, 130
            s1, s2 = 94, 80
            av, host = 8000, 8000
            facil_base = 12000
            photo, video = 4000, 7000
            print_base = 4000
            bus_base = 5000
            ins, mkt, team, legal = 1000, 5000, 8000, 3500
            cont_rate = 0.12
        else:  # Mid
            v12, v3 = 13000, 3500
            cat12, cat3 = 55, 30
            d1, d2 = 90, 90
            s1, s2 = 75, 60
            av, host = 5000, 5000
            facil_base = 8000
            photo, video = 2500, 4000
            print_base = 2500
            bus_base = 3000
            ins, mkt, team, legal = 500, 3000, 5000, 2000
            cont_rate = 0.10

        venue12 = f"IF({n}<=60,{v12}*0.85,IF({n}<=85,{v12},{v12}*1.15))"
        facil = f"IF({n}<=50,{facil_base},IF({n}<=75,{facil_base}*1.2,{facil_base}*1.4))"
        printing = f"{print_base}*0.7+{print_base}*0.3*{n}/70"
        bus = f"IF({n}<=50,{bus_base},{bus_base}*2)"

        subtotal = (
            f"({venue12})+{v3}"
            f"+{cat12}*{n}*2+{cat3}*{full}"
            f"+{d1}*{full}+{d2}*{full}"
            f"+{s1}*{full}+{s2}*{full}"
            f"+{av}+{host}+({facil})+{photo}+{video}"
            f"+({printing})+({bus})+{ins}+{mkt}+{team}+{legal}"
        )
        vat = f'IF(\'РАСХОДЫ\'!$C$4="Да",0,({subtotal})*0.15)'
        contingency = f"({subtotal})*{cont_rate}"

        return f"({subtotal})+({vat})+({contingency})"

    def _scenario_revenue(n, full_share, eb_share, lt, kn, p, s, t):
        """Build revenue formula for a specific scenario."""
        cs = "'ДОХОДЫ'!$C$2"  # creative share from param
        partner_tix = f"({p}*2+{s}*4+{t}*6)"

        cr_forum = f"ROUND({n}*{cs}*(1-{full_share}),0)"
        cr_full = f"ROUND({n}*{cs}*{full_share},0)"
        tech_forum = f"MAX(ROUND({n}*(1-{cs})*(1-{full_share}),0),0)"
        tech_full = f"MAX(ROUND({n}*(1-{cs})*{full_share},0)-{partner_tix},0)"

        cr_forum_rev = f"({cr_forum})*({eb_share}*290+(1-{eb_share})*377)"
        cr_full_rev = f"({cr_full})*({eb_share}*590+(1-{eb_share})*767)"
        tech_forum_rev = f"({tech_forum})*({eb_share}*490+(1-{eb_share})*637)"
        tech_full_rev = f"({tech_full})*({eb_share}*890+(1-{eb_share})*1157)"

        lt_rev = f"{lt}*500"
        kn_rev = f"{kn}*1500"
        partners = f"{p}*3000+{s}*8000+{t}*15000"

        return (f"{cr_forum_rev}+{cr_full_rev}"
                f"+{tech_forum_rev}+{tech_full_rev}"
                f"+{lt_rev}+{kn_rev}+{partners}")

    scenarios = [
        # (label, n, level, eb_share, full_share, lt, kn, p, s, t)
        ("Пессимист", 40, "High", 0.8, 0.4, 4, 0, 0, 0, 0),
        ("Базовый",   70, "Mid",  0.5, 0.6, 8, 1, 1, 1, 0),
        ("Оптимист", 100, "Low",  0.3, 0.7, 12, 2, 0, 1, 1),
    ]

    result_labels = ["Расходы", "Доходы", "Результат", "Маржа"]
    for i, label in enumerate(result_labels):
        r = res_start + i
        _set(ws, r, 1, label, border=THIN_BORDER, font=Font(bold=True))

    for col_idx, (label, n, level, eb, fs, lt, kn, p, s, t) in enumerate(
            scenarios):
        col = 2 + col_idx
        exp = _scenario_expense(n, level, fs, lt, kn, p, s, t)
        rev = _scenario_revenue(n, fs, eb, lt, kn, p, s, t)
        full_exp = f"=({exp})+0.04*({rev})"

        _set(ws, res_start, col, full_exp, border=THIN_BORDER,
             number_format=EUR_FMT)
        _set(ws, res_start + 1, col, f"={rev}", border=THIN_BORDER,
             number_format=EUR_FMT)
        _set(ws, res_start + 2, col,
             f"={get_column_letter(col)}{res_start + 1}"
             f"-{get_column_letter(col)}{res_start}",
             border=THIN_BORDER, number_format=EUR_FMT_NEG,
             font=Font(bold=True, size=12))
        _set(ws, res_start + 3, col,
             f"=IF({get_column_letter(col)}{res_start + 1}=0,0,"
             f"{get_column_letter(col)}{res_start + 2}"
             f"/{get_column_letter(col)}{res_start + 1})",
             border=THIN_BORDER, number_format='0.0%')

    # Conditional formatting on result row
    ws.conditional_formatting.add(
        f"B{res_start + 2}:D{res_start + 2}",
        CellIsRule(operator="greaterThanOrEqual", formula=["0"],
                   fill=green_fill))
    ws.conditional_formatting.add(
        f"B{res_start + 2}:D{res_start + 2}",
        CellIsRule(operator="lessThan", formula=["0"], fill=red_fill))

    # ========== Block 6: Investment summary (rows after scenarios) ==========
    inv_start = res_start + 6
    _set(ws, inv_start, 1, "ИНВЕСТИЦИОННАЯ СВОДКА (Risk/Reward)",
         font=HEADER_FONT, fill=HEADER_FILL)
    _style_range(ws, inv_start, 1, 4, font=HEADER_FONT, fill=HEADER_FILL)

    _set(ws, inv_start + 2, 1, "Maximum downside (пессимист)",
         font=Font(bold=True))
    _set(ws, inv_start + 2, 2, f"=B{res_start + 2}",
         number_format=EUR_FMT_NEG, font=Font(bold=True, color="FF0000"))

    _set(ws, inv_start + 3, 1, "Базовый результат", font=Font(bold=True))
    _set(ws, inv_start + 3, 2, f"=C{res_start + 2}",
         number_format=EUR_FMT_NEG)

    _set(ws, inv_start + 4, 1, "Base case маржа", font=Font(bold=True))
    _set(ws, inv_start + 4, 2, f"=C{res_start + 3}",
         number_format='0.0%')

    _set(ws, inv_start + 5, 1, "Upside маржа (оптимист)", font=Font(bold=True))
    _set(ws, inv_start + 5, 2, f"=D{res_start + 3}",
         number_format='0.0%')

    # % revenue from partners in base scenario
    _set(ws, inv_start + 7, 1, "Доля дохода от партнёров (базовый)",
         font=Font(bold=True))
    base_partner_rev = "1*3000+1*8000+0*15000"  # base: 1P+1S
    _set(ws, inv_start + 7, 2,
         f"=IF(C{res_start + 1}=0,0,({base_partner_rev})/C{res_start + 1})",
         number_format='0.0%')

    _set(ws, inv_start + 8, 1, "Доля дохода от билетов (базовый)",
         font=Font(bold=True))
    _set(ws, inv_start + 8, 2,
         f"=1-{get_column_letter(2)}{inv_start + 7}",
         number_format='0.0%')

    # Break-even without partners — find in break-even table where
    # 0-partner column turns green. Reference the matrix.
    _set(ws, inv_start + 10, 1,
         "Break-even без партнёров (0 партн., Mid)",
         font=Font(bold=True))
    _set(ws, inv_start + 10, 2,
         "См. столбец «0 партн.» в матрице выше",
         font=Font(italic=True, color="666666"))


def main():
    wb = openpyxl.Workbook()

    exp_info = _build_expenses(wb)
    rev_info = _build_revenue(wb, exp_info)
    _build_analysis(wb, exp_info, rev_info)

    output_path = "budget_model.xlsx"
    wb.save(output_path)
    print(f"Budget model saved to {output_path}")


if __name__ == "__main__":
    main()
